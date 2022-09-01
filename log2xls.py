from openpyxl import load_workbook
import os

# 读取x265 encoder log, 填入JCT-VC excel脚本

def check_log(encoder_type='origin'):
    """读取x265log,统计块划分的时间占比,写入表格encoder_time.xlsx"""
    if encoder_type == 'origin':
        modes = ["origin","origin_dpfast"]
    else:
        modes = ["anchor","anchor_dpfast"]
    seqs_info_path = "HEVC_Test_Sequence.txt"
    seqs_info_fp = open(seqs_info_path, 'r')
    seqs = []
    for line in seqs_info_fp:
        if line is None:
            break
        seqs.append(line.rstrip('\n'))

    for mode in modes:
        encoder_time = []
        bitrate = []
        Y_PSNR , U_PSNR , V_PSNR = [] , [] , []
        for seq_name in seqs:
            for qp in [22,27,32,37]:
                log_path = os.path.join(os.getcwd(),"log","encoder","enc_"+seq_name+'_'+mode+'_Q'+str(qp)+'.log') 
                with open(log_path,'r') as f:
                    lines = f.readlines()
                    command = False
                    for line in lines:
                        if command:
                            encoder_time.append(float(line.split(',')[2].strip(' ')))
                            bitrate.append(float(line.split(',')[4].strip(' ')))
                            Y_PSNR.append(float(line.split(',')[5].strip(' ')))
                            U_PSNR.append(float(line.split(',')[6].strip(' ')))
                            V_PSNR.append(float(line.split(',')[7].strip(' ')))
                        if line.split(',')[0] == 'Command':
                            command = True

                dec_log_path = os.path.join(os.getcwd(),"log","decoder","dec_"+seq_name+'_'+mode+'_Q'+str(qp)+'.log') 
                with open(dec_log_path,'r') as f:
                    lines = f.readlines()
                    command = False
                    for line in lines:
                        pass

                    
            wb = load_workbook(filename='JCTVC-L1100-' + modes[0] + '.xls')
            ws = wb['AI-Main']
            if mode == 'origin' or mode == 'anchor':
                index = 0
                for row in ws['D3:I82']:
                    for step,cell in enumerate(row):
                        if index % 4 == ((qp - 22) // 5):
                            if step == 0:
                                cell.value = bitrate[index]
                            elif step == 1:
                                cell.value = Y_PSNR[index]
                            elif step == 2:
                                cell.value = U_PSNR[index]
                            elif step == 3:
                                cell.value = V_PSNR[index]
                            elif step == 4:
                                cell.value = encoder_time[index]
                    index += 1
            elif mode == 'origin_dpfast' or mode == 'anchor_dpfast':
                index = 0
                for row in ws['L3:Q82']:
                    for step,cell in enumerate(row):
                        if index % 4 == ((qp - 22) // 5):
                            if step == 0:
                                cell.value = bitrate[index]
                            elif step == 1:
                                cell.value = Y_PSNR[index]
                            elif step == 2:
                                cell.value = U_PSNR[index]
                            elif step == 3:
                                cell.value = V_PSNR[index]
                            elif step == 4:
                                cell.value = encoder_time[index]
                    index += 1
            
            wb.save('encoder_time.xlsx')
            wb.close()

if __name__ == "__main__":
    check_log('origin') 
    # check_log('anchor')